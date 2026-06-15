"""SoS / rating data-integrity audit.

Flags where the engine's per-team strength inputs look unsupported by evidence, so an SoS
override or a form correction can be REVIEWED. It changes no data; it logs a single
'proposed' agent_run summarizing the flags (guardrail #1 — propose, don't apply; guardrail
#2 — never fabricate a replacement value).

Checks:
  A. SoS coverage & provenance — teams sitting on a bare confederation default (no evidence
     override), and whether team_form.notes actually carries provenance text.
  B. Ratings contradicted by observed WC results — teams whose prior_power was pulled
     hardest by the in-tournament re-rate (|post - prior|). That swing is real, sourced
     evidence the prior (form x SoS) is off; we cross-flag which sit on a default SoS.
  C. Form-source drift — DB team_form vs the workbook's Form_L20 (if present). Surfaces when
     the live data and the documented seed source disagree (so a re-seed would clobber).

    python scripts/audit_sos.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

import config
import lib.db as db
from engine.params import CONFED_SOS

SWING_FLAG = 4.0          # |post-prior| power swing that counts as a real contradiction
DOM_GA_PER_G = 0.6        # "stingy" threshold for the dominance/inflation heuristic
DOM_MIN_PLAYED = 6


def load(conn):
    rows = [dict(r) for r in conn.execute("""
        SELECT t.id, t.name, t.confederation,
               f.played, f.wins, f.draws, f.losses, f.gf, f.ga, f.sos, f.notes,
               p.prior_power, p.power, p.wc_games
        FROM teams t
        JOIN team_form f ON f.team_id = t.id
        JOIN power_ratings p ON p.team_id = t.id""")]
    for r in rows:
        r["default_sos"] = CONFED_SOS.get(r["confederation"])
        r["on_default"] = (r["default_sos"] is not None
                           and abs((r["sos"] or 0) - r["default_sos"]) < 1e-9)
        r["ga_per_g"] = (r["ga"] / r["played"]) if r["played"] else 0.0
        r["swing"] = round((r["power"] or 0) - (r["prior_power"] or 0), 1)
    return rows


def load_workbook_form():
    """Workbook Form_L20 played/W/D/L/GF/GA by team name, or None if unavailable."""
    candidates = [config.XLSX_PATH,
                  str(config.ROOT / "WorldCup2026_Analytics_Companion.xlsx"),
                  "WorldCup2026_Analytics_Companion.xlsx"]
    path = next((c for c in candidates if c and Path(c).is_file()), None)
    if not path:
        return None
    try:
        import warnings
        warnings.filterwarnings("ignore")
        import openpyxl
        fm = openpyxl.load_workbook(path, data_only=True)["Form_L20"]
    except Exception:
        return None
    out = {}
    for r in range(5, 53):
        nm = fm.cell(r, 2).value
        if nm:
            out[nm] = [fm.cell(r, c).value for c in (5, 6, 7, 8, 10, 11)]
    return out


def main():
    conn = db.connect()
    db.init_db(conn)
    rows = load(conn)
    by_name = {r["name"]: r for r in rows}
    n = len(rows)

    print("\nSoS / RATING DATA-INTEGRITY AUDIT\n")

    # --- A. SoS coverage & provenance ---
    on_default = [r for r in rows if r["on_default"]]
    notes_junk = sum(1 for r in rows
                     if (str(r["notes"]).strip().isdigit() if r["notes"] is not None else True))
    print("A. SoS coverage & provenance")
    print(f"   {len(on_default)}/{n} teams sit on a bare confederation-default SoS (no override).")
    prov = "MISSING — notes hold stray numbers, not provenance" if notes_junk > n // 2 \
        else f"{n - notes_junk}/{n} carry text"
    print(f"   team_form.notes provenance: {prov}.")
    print(f"   -> SoS overrides without sourced provenance violate the guardrail; the values")
    print(f"      may be fine, but they aren't auditable as written.\n")

    # --- B. ratings contradicted by observed results ---
    played = [r for r in rows if (r["wc_games"] or 0) > 0 and abs(r["swing"]) >= SWING_FLAG]
    played.sort(key=lambda r: -abs(r["swing"]))
    print(f"B. Ratings contradicted by observed WC results  (|post-prior| >= {SWING_FLAG})")
    if not played:
        print("   none yet.\n")
    else:
        print(f"   {'team':16s} {'conf':9s} {'prior':>6} {'post':>6} {'swing':>6}  SoS        verdict")
        for r in played:
            direction = "OVER-rated" if r["swing"] < 0 else "UNDER-rated"
            sos = f"{r['sos']:.2f}{'*def' if r['on_default'] else ' ovr'}"
            print(f"   {r['name'][:16]:16s} {r['confederation']:9s} "
                  f"{r['prior_power']:6.1f} {r['power']:6.1f} {r['swing']:+6.1f}  {sos:9s}  "
                  f"results say {direction}")
        ndef = sum(1 for r in played if r["on_default"])
        print(f"   -> {ndef}/{len(played)} of these sit on a default SoS: the model's worst calls")
        print(f"      cluster on un-tuned strength-of-schedule.\n")

    # --- B2. suspected inflation not yet tested by results ---
    dom = [r for r in rows if r["on_default"] and r["losses"] == 0
           and r["played"] >= DOM_MIN_PLAYED and r["ga_per_g"] <= DOM_GA_PER_G
           and abs(r["swing"]) < SWING_FLAG]
    if dom:
        print("B2. Suspected inflation (default SoS + unbeaten + stingy, not yet tested)")
        for r in sorted(dom, key=lambda r: -r["power"]):
            print(f"   {r['name'][:16]:16s} {r['confederation']:9s} rec {r['wins']}-{r['draws']}-{r['losses']} "
                  f"GA/g {r['ga_per_g']:.2f}  sos {r['sos']:.2f}  power {r['power']:.1f}  "
                  f"(dominant record on an un-discounted default)")
        print()

    # --- C. form-source drift vs workbook ---
    wbf = load_workbook_form()
    print("C. Form-source drift  (live DB team_form vs workbook Form_L20)")
    drift = []
    if wbf is None:
        print("   workbook not found — skipped.\n")
    else:
        keys = ["played", "wins", "draws", "losses", "gf", "ga"]
        for nm, w in wbf.items():
            d = by_name.get(nm)
            if d and any((w[i] or 0) != (d[keys[i]] or 0) for i in range(6)):
                drift.append(nm)
        print(f"   {len(drift)}/{len(wbf)} teams differ from the workbook source.")
        if len(drift) > len(wbf) // 2:
            print(f"   -> The live DB is NOT in sync with the workbook. Re-running")
            print(f"      seed_from_xlsx.py would OVERWRITE the engine's data. Do not re-seed")
            print(f"      until the intended source of truth is settled.")
        for nm in [t for t in ("Tunisia", "Sweden", "Switzerland") if t in drift]:
            d = by_name[nm]
            w = wbf[nm]
            print(f"     {nm}: workbook {w[:6]}  vs  DB [{d['played']},{d['wins']},{d['draws']},"
                  f"{d['losses']},{d['gf']},{d['ga']}]")
        print()

    # --- proposal (no data changed) ---
    db.log_run(conn, "sos_auditor", "flag_sos_outliers", {
        "on_default_sos": len(on_default),
        "notes_provenance_missing": notes_junk > n // 2,
        "results_contradicted": [
            {"team": r["name"], "swing": r["swing"], "on_default": r["on_default"]}
            for r in played],
        "suspected_inflation": [r["name"] for r in dom],
        "form_drift_vs_workbook": len(drift),
    }, status="proposed")
    conn.commit()
    conn.close()
    print("Logged a 'proposed' agent_run (sos_auditor). No data was changed — these are")
    print("flags for your review, not edits.")


if __name__ == "__main__":
    main()
