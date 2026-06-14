"""Seed the local DB from the existing Excel engine.

One-time bridge: reads Form_L20 + Tactics + Predictions from the workbook you already
built and validated, so you don't re-transcribe 48 teams. After this, the agents own
the data.

    python scripts/seed_from_xlsx.py [path/to/WorldCup2026_Analytics_Companion.xlsx]
"""
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import config
import lib.db as db
from engine.power import TeamForm, power


def slug(name: str) -> str:
    return name.strip().lower().replace(" ", "-").replace("'", "")


def find_xlsx(explicit: str | None) -> str:
    """Locate the workbook: explicit arg -> config -> cwd -> repo root -> Downloads."""
    candidates = [explicit, config.XLSX_PATH,
                  os.path.join(os.getcwd(), "WorldCup2026_Analytics_Companion.xlsx"),
                  str(Path(config.ROOT) / "WorldCup2026_Analytics_Companion.xlsx"),
                  str(Path.home() / "Downloads" / "WorldCup2026_Analytics_Companion.xlsx")]
    for c in candidates:
        if c and os.path.isfile(c):
            return c
    looked = "\n  ".join(str(c) for c in candidates if c)
    raise SystemExit(
        "Could not find WorldCup2026_Analytics_Companion.xlsx. Looked in:\n  "
        + looked
        + '\n\nPass the path explicitly, e.g.:\n'
          '  python scripts\\seed_from_xlsx.py '
          '"%USERPROFILE%\\Downloads\\WorldCup2026_Analytics_Companion.xlsx"'
    )


def main(xlsx_path: str | None = None):
    xlsx_path = find_xlsx(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    fm, tac, pred = wb["Form_L20"], wb["Tactics"], wb["Predictions"]

    # tactics lookup: H(8)=pressing, J(10)=pass_acc
    tactics = {}
    for r in range(5, 53):
        nm = tac.cell(row=r, column=2).value
        if nm:
            tactics[nm] = (tac.cell(row=r, column=8).value, tac.cell(row=r, column=10).value)

    conn = db.connect()
    db.init_db(conn)

    n = 0
    for r in range(5, 53):
        name = fm.cell(row=r, column=2).value
        if not name:
            continue
        tid = slug(name)
        press, pa = tactics.get(name, (6, 80))
        form = dict(
            played=fm.cell(row=r, column=5).value, wins=fm.cell(row=r, column=6).value,
            draws=fm.cell(row=r, column=7).value, losses=fm.cell(row=r, column=8).value,
            gf=fm.cell(row=r, column=10).value, ga=fm.cell(row=r, column=11).value,
            pass_acc=pa or 80, pressing=press or 6,
            sos=fm.cell(row=r, column=17).value, notes=fm.cell(row=r, column=16).value,
        )
        db.upsert_team(conn, tid, name, fm.cell(row=r, column=3).value, fm.cell(row=r, column=4).value)
        db.upsert_form(conn, tid, **form)

        p = power(TeamForm(form["played"], form["wins"], form["draws"], form["losses"],
                           form["gf"], form["ga"], form["pass_acc"], form["pressing"], form["sos"]))
        db.upsert_power(conn, tid, p, prior_power=p, wc_games=0, version=1)
        n += 1

    # matches from the Predictions tab (group stage). Q(17)/R(18) hold played scores.
    m = 0
    for r in range(6, 78):
        home = pred.cell(row=r, column=4).value
        away = pred.cell(row=r, column=5).value
        if not home or not away:
            continue
        hg, ag = pred.cell(row=r, column=17).value, pred.cell(row=r, column=18).value
        played = hg is not None and ag is not None
        d = pred.cell(row=r, column=3).value  # Date column
        kickoff = d.date().isoformat() if hasattr(d, "date") else (str(d) if d else None)
        db.upsert_match(conn, dict(
            id=f"g-{r}", stage="group", group_code=pred.cell(row=r, column=2).value,
            kickoff=kickoff, home_id=slug(home), away_id=slug(away),
            home_goals=hg if played else None, away_goals=ag if played else None,
            status="final" if played else "scheduled", source="seed:xlsx",
        ))
        m += 1

    conn.commit()
    conn.close()
    print(f"Seeded {n} teams + {m} matches into {config.DB_PATH}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else None)
